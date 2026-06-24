from flask import Flask, request, jsonify
import requests
import io
import zipfile
import re
import traceback
import base64
import os
import json
import ftplib
from datetime import datetime, timedelta, timezone

app = Flask(__name__)

DAYS_RU = {0: 'пн', 1: 'вт', 2: 'ср', 3: 'чт', 4: 'пт', 5: 'сб', 6: 'вс'}

HOLIDAYS = {
    '2026-05-01', '2026-05-02', '2026-05-03',
    '2026-05-09', '2026-05-10', '2026-05-11',
}

def get_tt_reference_from_ftp():
    try:
        host = os.environ.get('FTP_HOST', '185.123.193.181')
        port = int(os.environ.get('FTP_PORT', 38021))
        user = os.environ.get('FTP_USER', 'контроль')
        password = os.environ.get('FTP_PASS', '147258369')

        ftp = ftplib.FTP()
        ftp.connect(host, port, timeout=30)
        ftp.login(user, password)
        ftp.set_pasv(True)

        buffer = io.BytesIO()
        ftp.retrbinary('RETR /OData/OData_83_ТТ.txt', buffer.write)
        ftp.quit()

        buffer.seek(0)
        raw = buffer.read()
        text = raw.decode('utf-8-sig', errors='ignore')
        import re as _re
        text = text.replace('\r\n', ' ').replace('\r', ' ')
        text = _re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]', '', text)
        try:
            return json.loads(text)
        except Exception:
            text2 = ''.join(c for c in text if c.isprintable() or c in '\t\n\r')
            return json.loads(text2)
    except Exception as e:
        print(f'FTP error: {e}')
        return None

def get_sheet_name_for_mode(date_mode='yesterday'):
    MSK = timezone(timedelta(hours=3))
    now = datetime.now(MSK)

    if date_mode == 'today':
        d = now
        if d.weekday() >= 5 or d.strftime('%Y-%m-%d') in HOLIDAYS:
            return None
    else:
        d = now - timedelta(days=1)
        while d.weekday() >= 5 or d.strftime('%Y-%m-%d') in HOLIDAYS:
            d = d - timedelta(days=1)

    day = d.day
    weekday = DAYS_RU[d.weekday()]
    return f"{day:02d} {weekday}"

def get_yesterday_sheet_name():
    return get_sheet_name_for_mode('yesterday')

def load_workbook_safe(content):
    zin = zipfile.ZipFile(io.BytesIO(content))
    zout_buffer = io.BytesIO()
    zout = zipfile.ZipFile(zout_buffer, 'w', zipfile.ZIP_DEFLATED)
    for item in zin.infolist():
        data = zin.read(item.filename)
        if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
            try:
                xml = data.decode('utf-8')
                xml = re.sub(r'<dataValidations[^>]*>.*?</dataValidations>', '', xml, flags=re.DOTALL)
                xml = re.sub(r'<dataValidation[^/]*/>', '', xml)
                data = xml.encode('utf-8')
            except Exception:
                pass
        zout.writestr(item, data)
    zout.close()
    zout_buffer.seek(0)
    return zout_buffer.read()

def get_sheet_day(sheet_name):
    try:
        return int(sheet_name.split()[0])
    except:
        return None

def get_prev_sheets_7days(wb, today_sheet):
    """Возвращает все листы за последние 7 календарных дней (по возрастанию дня)."""
    today_day = get_sheet_day(today_sheet)
    if today_day is None:
        return []

    candidates = []
    for name in wb.sheetnames:
        if name in ('Контроль', 'КОДЫ ТТ'):
            continue
        day = get_sheet_day(name)
        if day is not None and (today_day - 7) <= day < today_day:
            candidates.append((day, name))

    candidates.sort(key=lambda x: x[0])
    return [name for _, name in candidates]

def get_next_sheet_same_weekday(wb, today_sheet):
    today_day = get_sheet_day(today_sheet)
    if today_day is None:
        return None
    parts = today_sheet.split()
    if len(parts) < 2:
        return None
    today_weekday = parts[1]

    candidates = []
    for name in wb.sheetnames:
        if name in ('Контроль', 'КОДЫ ТТ'):
            continue
        name_parts = name.split()
        if len(name_parts) < 2:
            continue
        day = get_sheet_day(name)
        weekday = name_parts[1]
        if day is not None and day > today_day and weekday == today_weekday:
            candidates.append((day, name))

    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]

def find_columns(ws):
    tt_col = sum_col = plan_col = dev_col = otklonenie_col = header_row = None
    for row in ws.iter_rows(max_row=5):
        for cell in row:
            v = cell.value
            if v == 'Код ТТ': tt_col = cell.column
            if v == 'Сумма заявки': sum_col = cell.column; header_row = cell.row
            if v == 'План сумма': plan_col = cell.column
            if v == 'Отклонения дня': dev_col = cell.column
            if v == 'Отклонение': otklonenie_col = cell.column
        if header_row:
            break
    return tt_col, sum_col, plan_col, dev_col, otklonenie_col, header_row

def find_itogo_row(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and str(cell.value).strip() == 'Итого' and cell.column == 1:
                return cell.row
    return None

def build_deviation_formula(prev_sheets, tt_col_letter, row_num, max_rows=300):
    """Формула VLOOKUP суммирующая Отклонение за последние 7 дней."""
    if not prev_sheets:
        return None
    parts = []
    for sheet in prev_sheets:
        parts.append(
            f"IFERROR(VLOOKUP(${tt_col_letter}{row_num},'{sheet}'!$B$1:$R${max_rows},17,FALSE),0)"
        )
    return '=' + '+'.join(parts)

@app.route('/debug-date', methods=['GET'])
def debug_date():
    MSK = timezone(timedelta(hours=3))
    now = datetime.now(MSK)
    return jsonify({
        'now_msk': now.isoformat(),
        'weekday': now.weekday(),
        'today_sheet': get_sheet_name_for_mode('today'),
        'yesterday_sheet': get_sheet_name_for_mode('yesterday'),
        'holidays': sorted(HOLIDAYS)
    })

@app.route('/update-excel', methods=['POST'])
def update_excel():
    try:
        data = request.json
        webhook = data.get('webhook')
        file_id = data.get('file_id')
        updates = data.get('updates')
        date_mode = data.get('date_mode', 'yesterday')

        MSK = timezone(timedelta(hours=3))
        now_msk = datetime.now(MSK)
        today_sheet = get_sheet_name_for_mode(date_mode)
        debug_info = {
            'date_mode_received': date_mode,
            'now_msk': now_msk.isoformat(),
            'weekday': now_msk.weekday(),
            'today_sheet_calculated': today_sheet
        }
        if today_sheet is None:
            return jsonify({'status': 'skipped', 'message': 'Сегодня выходной или праздник, запись пропущена', 'debug': debug_info})

        r = requests.get(f'{webhook}/disk.file.get.json?id={file_id}')
        file_info = r.json()
        download_url = file_info['result']['DOWNLOAD_URL']
        file_name = file_info['result']['NAME']

        r = requests.get(download_url)
        clean_content = load_workbook_safe(r.content)

        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(clean_content))

        if today_sheet not in wb.sheetnames:
            return jsonify({'status': 'error', 'message': f'Вкладка {today_sheet} не найдена', 'sheets': wb.sheetnames}), 400

        ws = wb[today_sheet]
        tt_col, sum_col, plan_col, dev_col, otklonenie_col, header_row = find_columns(ws)

        if not header_row or not sum_col or not tt_col:
            return jsonify({'status': 'error', 'message': 'Колонки не найдены'}), 400

        updates_map = {str(u.get('tt_code')).strip(): u.get('fact') for u in updates}

        # Шаг 1 — записываем факт и собираем найденные ТТ
        updated = 0
        found_codes = set()
        for row in ws.iter_rows(min_row=header_row + 1):
            tt = str(row[tt_col - 1].value or '').strip().replace('\xa0', '').replace(' ', '')
            if not tt:
                continue
            found_codes.add(tt)
            if tt in updates_map:
                row[sum_col - 1].value = round(float(updates_map[tt]), 2)
                updated += 1

        # Шаг 2 — записываем формулу "Отклонения дня" в следующий лист
        dev_updated = 0
        next_sheet = get_next_sheet_same_weekday(wb, today_sheet)

        if next_sheet and next_sheet in wb.sheetnames and dev_col:
            ws_next = wb[next_sheet]
            tt_col_n, _, _, dev_col_n, _, header_row_n = find_columns(ws_next)
            prev_sheets_7 = get_prev_sheets_7days(wb, next_sheet)

            if dev_col_n and header_row_n and tt_col_n and prev_sheets_7:
                tt_col_letter = chr(64 + tt_col_n)
                for row in ws_next.iter_rows(min_row=header_row_n + 1):
                    tt = str(row[tt_col_n - 1].value or '').strip().replace('\xa0', '').replace(' ', '')
                    if not tt or not tt.startswith('T'):
                        continue
                    formula = build_deviation_formula(
                        prev_sheets_7, tt_col_letter, row[0].row
                    )
                    if formula:
                        row[dev_col_n - 1].value = formula
                        dev_updated += 1

        # Шаг 3 — определяем region_codes из всех плановых листов файла
        tt_reference = get_tt_reference_from_ftp()

        region_codes = set()
        skip_sheets = {'Контроль', 'КОДЫ ТТ', today_sheet}
        for sheet_name in wb.sheetnames:
            if sheet_name in skip_sheets:
                continue
            ws_plan = wb[sheet_name]
            tt_col_p, _, _, _, _, header_row_p = find_columns(ws_plan)
            if not tt_col_p or not header_row_p:
                continue
            for row in ws_plan.iter_rows(min_row=header_row_p + 1):
                tt = str(row[tt_col_p - 1].value or '').strip().replace('\xa0', '').replace(' ', '')
                if tt and tt.startswith('T') and len(tt) == 5:
                    region_codes.add(tt)

        # Шаг 4 — добавляем внеплановые ТТ
        unplanned_added = 0
        debug_not_found = []

        tt_info_map = {}
        if tt_reference:
            for row in tt_reference:
                code = str(row.get('КодТорговойТочки') or '').strip()
                if code:
                    tt_info_map[code] = {
                        'name': str(row.get('НаименованиеТТ') or '').strip().replace(';', ','),
                        'route': str(row.get('МаршрутТТ') or '').strip()
                    }

        itogo_row = find_itogo_row(ws)

        last_tt_row = header_row
        first_empty_before_summary = None

        if itogo_row:
            for r in range(itogo_row - 1, header_row, -1):
                tt = str(ws.cell(row=r, column=tt_col).value or '').strip()
                if tt and tt.startswith('T') and len(tt) == 5:
                    last_tt_row = r
                    break
            first_empty_before_summary = last_tt_row + 1

        first_summary_row = first_empty_before_summary

        unplanned_to_add = []
        for tt_code, fact in updates_map.items():
            if tt_code not in found_codes:
                debug_not_found.append(tt_code)
                # Временно отключено до получения файла с РМ от Вадима
                # if fact and float(fact) != 0:
                #     if not region_codes or tt_code in region_codes:
                #         unplanned_to_add.append((tt_code, fact))

        if first_summary_row and itogo_row:
            needed_rows = len(unplanned_to_add) + 1
            available_rows = itogo_row - first_summary_row
            if needed_rows > available_rows:
                insert_count = needed_rows - available_rows
                ws.insert_rows(first_summary_row, amount=insert_count)

        current_row = last_tt_row + 1
        for tt_code, fact in unplanned_to_add:
            ws.cell(row=current_row, column=tt_col).value = tt_code
            ws.cell(row=current_row, column=sum_col).value = round(float(fact), 2)

            info = tt_info_map.get(tt_code, {})
            tt_cell = f"{chr(64 + tt_col)}{current_row}"

            if info.get('route'):
                ws.cell(row=current_row, column=3).value = info['route']
            else:
                ws.cell(row=current_row, column=3).value = f'=IFERROR(VLOOKUP({tt_cell},\'КОДЫ ТТ\'!$A:$G,2,FALSE),"")'

            ws.cell(row=current_row, column=4).value = f'=IFERROR(VLOOKUP({tt_cell},\'КОДЫ ТТ\'!$A:$G,3,FALSE),"")'
            ws.cell(row=current_row, column=5).value = f'=IFERROR(VLOOKUP({tt_cell},\'КОДЫ ТТ\'!$A:$G,4,FALSE),"")'

            if info.get('name'):
                ws.cell(row=current_row, column=6).value = info['name']
            else:
                ws.cell(row=current_row, column=6).value = f'=IFERROR(VLOOKUP({tt_cell},\'КОДЫ ТТ\'!$A:$G,5,FALSE),"")'

            ws.cell(row=current_row, column=10).value = f'=IFERROR(VLOOKUP({tt_cell},\'КОДЫ ТТ\'!$A:$G,6,FALSE),"")'
            ws.cell(row=current_row, column=11).value = f'=IFERROR(VLOOKUP({tt_cell},\'КОДЫ ТТ\'!$A:$G,7,FALSE),"")'
            current_row += 1
            unplanned_added += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        file_content_b64 = base64.b64encode(output.read()).decode('utf-8')

        upload_r = requests.post(
            f'{webhook}/disk.file.uploadversion.json',
            json={'id': file_id, 'fileContent': ['file.xlsx', file_content_b64]}
        )

        return jsonify({
            'status': 'ok',
            'sheet': today_sheet,
            'next_sheet': next_sheet,
            'updated_fact': updated,
            'updated_deviation': dev_updated,
            'unplanned_added': unplanned_added,
            'debug_not_found': debug_not_found,
            'ftp_reference_loaded': tt_reference is not None,
            'debug': debug_info,
            'result': upload_r.json()
        })

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e), 'trace': traceback.format_exc()}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=10000)
